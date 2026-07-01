from io import BytesIO

from openpyxl import Workbook


class ExcelTemplateService:
    # ===========================================
    # STUDENT TEMPLATE
    # ===========================================

    def student_template(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Students"

        ws.append(
            [
                "first_name",
                "last_name",
                "username",
                "email",
                "gender",
                "date_of_birth",
                "admission_date",
                "class_id",
            ]
        )

        ws.append(
            [
                "John",
                "Doe",
                "john001",
                "john@example.com",
                "male",
                "2012-04-15",
                "2024-09-10",
                "CLASS_UUID_HERE",
            ]
        )

        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)

        return stream

    # ===========================================
    # TEACHER TEMPLATE
    # ===========================================

    def teacher_template(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Teachers"

        ws.append(
            [
                "first_name",
                "last_name",
                "username",
                "email",
                "qualification",
                "specialization",
                "hire_date",
                "class_id",
            ]
        )

        ws.append(
            [
                "Sarah",
                "James",
                "sarahj",
                "sarah@example.com",
                "B.Ed",
                "Mathematics",
                "2024-01-12",
                "CLASS_UUID_HERE",
            ]
        )

        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)

        return stream

    # ===========================================
    # PARENT TEMPLATE
    # ===========================================

    def parent_template(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Parents"

        ws.append(
            [
                "first_name",
                "last_name",
                "username",
                "email",
                "occupation",
                "phone",
            ]
        )

        ws.append(
            [
                "David",
                "Johnson",
                "davidp",
                "david@example.com",
                "Engineer",
                "+2348012345678",
            ]
        )

        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)

        return stream


excel_template_service = ExcelTemplateService()
