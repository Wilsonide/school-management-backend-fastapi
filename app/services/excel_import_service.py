from io import BytesIO

from openpyxl import load_workbook


class ExcelImportService:
    def read_students(self, file):
        workbook = load_workbook(
            BytesIO(file),
        )

        sheet = workbook.active

        students = []

        for row in sheet.iter_rows(
            min_row=2,
            values_only=True,
        ):
            students.append(
                {
                    "first_name": row[0],
                    "last_name": row[1],
                    "email": row[2],
                    "username": row[3],
                    "gender": row[4],
                    "date_of_birth": row[5],
                    "admission_date": row[6],
                    "class_id": row[7],
                }
            )

        return students
        # =====================================================

    # READ TEACHERS
    # =====================================================

    def read_teachers(self, file):
        workbook = load_workbook(
            BytesIO(file),
        )

        sheet = workbook.active

        teachers = []

        for row in sheet.iter_rows(
            min_row=2,
            values_only=True,
        ):
            teachers.append(
                {
                    "first_name": row[0],
                    "last_name": row[1],
                    "email": row[2],
                    "username": row[3],
                    "qualification": row[4],
                    "specialization": row[5],
                    "hire_date": row[6],
                    "class_id": row[7],
                }
            )

        return teachers


excel_import_service = ExcelImportService()
