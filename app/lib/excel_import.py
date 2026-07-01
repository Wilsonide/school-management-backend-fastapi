from io import BytesIO

from fastapi import HTTPException, UploadFile, status
from openpyxl import load_workbook


class ExcelImport:
    @staticmethod
    async def read(upload: UploadFile):
        """
        Read an uploaded Excel file and return
        the active worksheet.
        """

        if not upload.filename.endswith(".xlsx"):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Only .xlsx files are supported.",
            )

        contents = await upload.read()

        workbook = load_workbook(
            BytesIO(contents),
            data_only=True,
        )

        return workbook.active

    @staticmethod
    def headers(sheet):
        """
        Returns the first row as headers.
        """

        return [
            str(cell.value).strip() if cell.value is not None else ""
            for cell in sheet[1]
        ]

    @staticmethod
    def validate_headers(
        sheet,
        expected_headers,
    ):
        headers = ExcelImport.headers(sheet)

        if headers != expected_headers:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail={
                    "message": "Invalid Excel template.",
                    "expected": expected_headers,
                    "received": headers,
                },
            )

    @staticmethod
    def rows(sheet):
        """Returns all data rows excluding header."""
        yield from sheet.iter_rows(
            min_row=2,
            values_only=True,
        )
